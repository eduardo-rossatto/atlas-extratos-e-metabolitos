"""
Ficha Técnica Comercial B2B — BIOTROP
Versão web (Streamlit)
"""

import streamlit as st
import streamlit.components.v1 as components
import io, os, re, zipfile
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.drawing.image import Image as XLImage

# ── Página ────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ATLAS - E&M",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Tema Groq Light × BIOTROP Green ───────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

/* ═══════════════════════════════════════════════════════
   ATLAS — Groq Light × BIOTROP Green
   #FAFAF8  bg principal (warm off-white)
   #F3F3EE  bg secundário
   #E8E8DE  bg elevado / hover
   #D0D0C8  bordas
   #52b788  accent primário (verde BIOTROP)
   #40916c  accent dark
   #2D6A4F  accent deep
   #2D2F33  texto primário
   #69695D  texto secundário
   #9C9C90  texto muted
   ═══════════════════════════════════════════════════════ */

html, body, input, textarea, select,
label, p, h1, h2, h3, h4, h5, h6,
.stMarkdown, .stTextInput, .stSelectbox,
.stTextArea, .stButton > button > div > p {
    font-family: 'Space Grotesk', sans-serif !important;
}

/* ── Reset global ─────────────────────────────────────── */
html, body,
[data-testid="stAppViewContainer"],
[data-testid="stApp"],
.main, .block-container {
    background-color: #FAFAF8 !important;
    color: #2D2F33 !important;
}

[data-testid="stHeader"],
[data-testid="stToolbar"] {
    background-color: #FAFAF8 !important;
    border-bottom: 1px solid #E8E8DE !important;
}

[data-testid="stSidebar"] {
    background-color: #F3F3EE !important;
}

section[data-testid="stVerticalBlock"],
div[data-testid="stVerticalBlock"] {
    background-color: transparent !important;
}

/* ── Abas ─────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    background-color: #F3F3EE !important;
    border-radius: 1000px !important;
    padding: 4px 6px !important;
    gap: 2px !important;
    overflow-x: auto !important;
    flex-wrap: nowrap !important;
    border: 1px solid #D0D0C8 !important;
}

.stTabs [data-baseweb="tab"] {
    background-color: transparent !important;
    color: #9C9C90 !important;
    font-weight: 500 !important;
    font-size: 12px !important;
    border-radius: 1000px !important;
    padding: 6px 14px !important;
    white-space: nowrap !important;
    border: none !important;
    transition: all 0.2s ease !important;
    letter-spacing: 0.02em !important;
}

.stTabs [data-baseweb="tab"]:hover {
    background-color: #E8E8DE !important;
    color: #2D2F33 !important;
}

.stTabs [aria-selected="true"] {
    background-color: #52b788 !important;
    color: #FFFFFF !important;
    box-shadow: 0 2px 10px rgba(82,183,136,0.3) !important;
}

.stTabs [data-baseweb="tab-panel"] {
    background-color: #FAFAF8 !important;
    padding-top: 20px !important;
}

/* ── Inputs, textareas, selectbox ─────────────────────── */
input, textarea, select,
[data-baseweb="input"] input,
[data-baseweb="textarea"] textarea,
[data-baseweb="select"] div {
    background-color: #FFFFFF !important;
    color: #2D2F33 !important;
    border: 1px solid #D0D0C8 !important;
    border-radius: 10px !important;
    font-family: 'Space Grotesk', sans-serif !important;
}

[data-baseweb="input"],
[data-baseweb="textarea"],
[data-baseweb="base-input"] {
    background-color: #FFFFFF !important;
    border-color: #D0D0C8 !important;
    border-radius: 10px !important;
}

[data-baseweb="input"]:focus-within,
[data-baseweb="textarea"]:focus-within,
[data-baseweb="base-input"]:focus-within {
    background-color: #FFFFFF !important;
    border-color: #52b788 !important;
    box-shadow: 0 0 0 2px rgba(82,183,136,0.2) !important;
}

input:focus, textarea:focus {
    background-color: #FFFFFF !important;
    color: #2D2F33 !important;
    border-color: #52b788 !important;
    box-shadow: 0 0 0 2px rgba(82,183,136,0.2) !important;
    outline: none !important;
}

[data-baseweb="select"],
[data-baseweb="select"] > div,
[data-baseweb="select"] div,
[data-baseweb="select"] span,
[data-baseweb="select"] input {
    background-color: #FFFFFF !important;
    border-color: #D0D0C8 !important;
    color: #2D2F33 !important;
    border-radius: 10px !important;
}

[data-baseweb="select"]:focus-within > div {
    border-color: #52b788 !important;
    box-shadow: 0 0 0 2px rgba(82,183,136,0.2) !important;
}

[data-baseweb="popover"],
[data-baseweb="popover"] * {
    background-color: #FFFFFF !important;
    color: #2D2F33 !important;
}

[data-baseweb="popover"] [role="listbox"],
[data-baseweb="menu"] {
    background-color: #FFFFFF !important;
    border: 1px solid #D0D0C8 !important;
    border-radius: 12px !important;
    box-shadow: 0 4px 20px rgba(0,0,0,0.08) !important;
}

[data-baseweb="menu"] li,
[data-baseweb="option"],
[role="option"] {
    background-color: #FFFFFF !important;
    color: #2D2F33 !important;
}

[data-baseweb="menu"] li:hover,
[data-baseweb="option"]:hover,
[role="option"]:hover {
    background-color: #F3F3EE !important;
    color: #40916c !important;
}

/* ── Labels e textos ──────────────────────────────────── */
label, p, span, div,
[data-testid="stMarkdownContainer"] p,
[data-testid="stCaptionContainer"] {
    color: #2D2F33 !important;
}

h1, h2, h3, h4 {
    color: #2D2F33 !important;
    font-weight: 400 !important;
    font-family: 'Space Grotesk', sans-serif !important;
}

.biotrop-header h1 {
    font-weight: 700 !important;
}

/* ── Checkbox ─────────────────────────────────────────── */
[data-testid="stCheckbox"] label { color: #69695D !important; }
[data-testid="stCheckbox"] span[aria-checked="true"] {
    background-color: #52b788 !important;
    border-color: #52b788 !important;
}

/* ── Radio ────────────────────────────────────────────── */
[data-testid="stRadio"] label { color: #69695D !important; }

/* ── Divisor ──────────────────────────────────────────── */
hr { border-color: #E8E8DE !important; }

/* ── Botões ───────────────────────────────────────────── */
html body div[data-testid="stButton"] > button,
html body div[data-testid="stDownloadButton"] > button,
html body div[data-testid="stButton"] > button:focus,
html body div[data-testid="stDownloadButton"] > button:focus,
html body div[data-testid="stButton"] > button:active,
html body div[data-testid="stDownloadButton"] > button:active {
    background: #52b788 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 1000px !important;
    font-weight: 500 !important;
    font-size: 14px !important;
    padding: 10px 24px !important;
    box-shadow: 0 2px 10px rgba(82,183,136,0.25) !important;
    transition: all 0.2s ease !important;
    letter-spacing: 0.02em !important;
}

html body div[data-testid="stButton"] > button > p,
html body div[data-testid="stDownloadButton"] > button > p,
html body div[data-testid="stButton"] > button *,
html body div[data-testid="stDownloadButton"] > button * {
    color: #FFFFFF !important;
    font-weight: 500 !important;
}

html body div[data-testid="stButton"] > button:hover,
html body div[data-testid="stDownloadButton"] > button:hover {
    background: #40916c !important;
    box-shadow: 0 4px 16px rgba(82,183,136,0.4) !important;
    transform: scale(1.02) !important;
}

/* ── File uploader ────────────────────────────────────── */
/* Nuclear: sobrescreve variáveis CSS do Streamlit no uploader */
[data-testid="stFileUploader"] {
    --background-color: #FFFFFF;
    --secondary-background-color: #F3F3EE;
    --text-color: #2D2F33;
}

[data-testid="stFileUploader"],
[data-testid="stFileUploader"] div,
[data-testid="stFileUploader"] div div,
[data-testid="stFileUploader"] div div div,
[data-testid="stFileUploader"] section,
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] p {
    background-color: #FFFFFF !important;
    color: #2D2F33 !important;
}

[data-testid="stFileUploadDropzone"] {
    background-color: #FFFFFF !important;
    border: 1.5px dashed #D0D0C8 !important;
    border-radius: 12px !important;
    color: #2D2F33 !important;
}

[data-testid="stFileUploadDropzone"]:hover {
    border-color: #52b788 !important;
    background-color: #F3F3EE !important;
}

/* Botão "Browse files" */
[data-testid="stFileUploader"] button {
    background-color: #F3F3EE !important;
    color: #2D2F33 !important;
    border: 1px solid #D0D0C8 !important;
    border-radius: 1000px !important;
}

/* ── Chip do arquivo após upload ──────────────────────── */
[data-testid="stFileUploaderFile"],
[data-testid="stFileUploaderFile"] *,
[data-testid="stFileUploaderFileName"],
[data-testid="stFileUploaderFileName"] *,
[data-testid="stUploadedFile"],
[data-testid="stUploadedFile"] *,
.stUploadedFile,
.stUploadedFile *,
[class*="uploadedFile"],
[class*="uploadedFile"] *,
[class*="UploadedFile"],
[class*="UploadedFile"] * {
    background-color: #F3F3EE !important;
    color: #2D2F33 !important;
    border-color: #D0D0C8 !important;
}

/* Botão X de remover o arquivo */
[data-testid="stFileUploaderFile"] button,
[data-testid="stUploadedFile"] button,
.stUploadedFile button {
    background-color: transparent !important;
    color: #69695D !important;
}

/* ── Checkboxes centralizados ─────────────────────────── */
[data-testid="stCheckbox"] {
    display: flex !important;
    justify-content: center !important;
}
[data-testid="stCheckbox"] > label {
    display: flex !important;
    justify-content: center !important;
    align-items: center !important;
    width: 100% !important;
}
[data-testid="stCheckbox"] > label > div:first-child {
    margin: 0 auto !important;
}

/* ── Ocultar seletor de tema ──────────────────────────── */
[data-testid="stToolbarActions"] { display: none !important; }

/* ── Alertas ──────────────────────────────────────────── */
[data-testid="stAlert"] {
    background-color: #F3F3EE !important;
    border-color: #52b788 !important;
    border-radius: 12px !important;
    color: #2D2F33 !important;
}

/* ── Expander ─────────────────────────────────────────── */
[data-testid="stExpander"] {
    background-color: #FFFFFF !important;
    border: 1px solid #D0D0C8 !important;
    border-radius: 16px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
}

/* ── Componentes customizados ─────────────────────────── */
.biotrop-header {
    background-color: #FFFFFF !important;
    padding: 20px 28px !important;
    border-radius: 20px !important;
    margin-bottom: 20px !important;
    border: 1px solid #D0D0C8 !important;
    box-shadow: 0 2px 12px rgba(0,0,0,0.06) !important;
    color: #2D2F33 !important;
}

.biotrop-header h1 {
    font-weight: 700 !important;
    color: #2D2F33 !important;
}

.biotrop-header h2,
.biotrop-header p {
    color: inherit !important;
}

.sec-header {
    color: #40916c;
    padding-bottom: 6px;
    margin: 20px 0 10px 0;
    font-size: 14px;
    font-weight: 500;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    border-bottom: 1px solid #E8E8DE;
}

.step-guide {
    background-color: #F3F3EE;
    border-left: 2px solid #52b788;
    padding: 10px 16px;
    border-radius: 0 10px 10px 0;
    margin-bottom: 16px;
    color: #2D2F33;
    font-size: 13px;
    border-top: 1px solid #E8E8DE;
    border-right: 1px solid #E8E8DE;
    border-bottom: 1px solid #E8E8DE;
}

.next-hint {
    background-color: #F3F3EE;
    color: #40916c;
    padding: 10px 18px;
    border-radius: 1000px;
    margin-top: 20px;
    font-size: 13px;
    font-weight: 500;
    border: 1px solid #D0D0C8;
}
</style>
""", unsafe_allow_html=True)

# ── Constantes ────────────────────────────────────────────────────────────────
UNIDADES = ["End/ml", "End/L", "UFC/ml", "UFC/L"]

DOCS_INFO = [
    ("14. Bula",                          "C255", "H255", "Bula"),
    ("15. Rótulo",                        "C258", "H258", "Rotulo"),
    ("16. Certificado de Registro",       "C261", "H261", "Certificado de Registro"),
    ("17. Ficha de Emergência",           "C264", "H264", "Ficha de Emergencia"),
    ("18. FDS",                           "C267", "H267", "FDS"),
    ("18. SDS",                           "C270", "H270", "SDS"),
    ("Estudos de Estabilidade Acelerada", "",     "",     "Estabilidade Acelerada"),
    ("Estudos de Longa Duração",          "",     "",     "Longa Duracao"),
]

TEMPLATE_PATH = "template_extratos.xlsx"

ABAS = [
    "① Identificação",
    "② Concentração",
    "③ Classificação",
    "④ Recomendações",
    "⑤ Qualidade",
    "⑥ CoA / Estabilidade",
    "⑦ Documentos",
    "⑧ Embalagens",
    "⑨ Fotos",
    "⑩ Assinatura",
]

# ── Extração automática via PDF da Bula ───────────────────────────────────────
def extrair_texto_pdf(uploaded_file) -> str:
    import pdfplumber
    uploaded_file.seek(0)
    pages_text = []
    with pdfplumber.open(io.BytesIO(uploaded_file.read())) as pdf:
        for p in pdf.pages:
            parts = []

            # Texto corrido compacto (sem layout para não desperdiçar caracteres)
            text = p.extract_text() or ""
            if text.strip():
                parts.append(text)

            # Tabelas com separadores explícitos de coluna (preserva estrutura sem espaços)
            for table in (p.extract_tables() or []):
                rows = []
                for row in table:
                    cells = [str(c).strip().replace("\n", " ") if c else "" for c in row]
                    rows.append(" | ".join(cells))
                if rows:
                    parts.append("---TABELA---\n" + "\n".join(rows))

            if parts:
                pages_text.append("\n".join(parts))

    return "\n\n".join(pages_text)


def extrair_texto_excel(uploaded_file) -> str:
    import openpyxl
    uploaded_file.seek(0)
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
    linhas = []
    for sheet in wb.worksheets:
        linhas.append(f"[Aba: {sheet.title}]")
        for row in sheet.iter_rows():
            valores = [str(cell.value).strip() for cell in row if cell.value is not None]
            if valores:
                linhas.append(" | ".join(valores))
    return "\n".join(linhas)


def _extrair_tabelas_recomendacoes(uploaded_file) -> list:
    import pdfplumber
    rec_rows = []
    uploaded_file.seek(0)
    cabecalhos_alvo = {"cultura", "dose", "calda", "alvos", "aplicac", "interv",
                       "n°", "nº", "época", "volume", "produto"}
    with pdfplumber.open(io.BytesIO(uploaded_file.read())) as pdf:
        for page in pdf.pages:
            for tabela in (page.extract_tables() or []):
                if not tabela or len(tabela) < 2:
                    continue
                cabecalho = [str(c or "").lower().strip() for c in tabela[0]]
                if not any(any(alvo in c for alvo in cabecalhos_alvo) for c in cabecalho):
                    continue
                def col(keywords):
                    for i, c in enumerate(cabecalho):
                        if any(k in c for k in keywords):
                            return i
                    return None
                i_cult = col(["cultura", "crop"])
                i_alv  = col(["alvo", "praga", "doen", "target"])
                i_dose = col(["dose"])
                i_vol  = col(["calda", "volume", "vol"])
                i_num  = col(["n°", "nº", "aplic", "época", "interv", "número"])

                # Unidade do volume no cabeçalho (ex: "Volume (L/ha)" → "L/ha")
                import re as _re_vol
                _vol_unit = ""
                if i_vol is not None:
                    _h = cabecalho[i_vol]
                    _m = _re_vol.search(r'\(([^)]+)\)', _h)
                    if _m:
                        _vol_unit = _m.group(1).strip()
                    else:
                        _m2 = _re_vol.search(r'(m?[lL]/\S+)', _h)
                        if _m2:
                            _vol_unit = _m2.group(1)

                for linha in tabela[1:]:
                    def v(i):
                        return str(linha[i] or "").strip() if i is not None and i < len(linha) else ""
                    cult = v(i_cult)
                    if not cult:
                        continue
                    _vol = v(i_vol)
                    # Se a célula tem só número e a unidade está no cabeçalho, anexar
                    if _vol and _vol_unit and _re_vol.fullmatch(r'[\d\.,\s]+', _vol):
                        _vol = f"{_vol} {_vol_unit}"
                    rec_rows.append(dict(
                        cultura=cult, alvos=v(i_alv),
                        dose=v(i_dose), volume=_vol, numero=v(i_num)
                    ))
    return rec_rows


def _extrair_dados_bula_regex(texto: str) -> dict:
    import re

    def buscar(padroes, texto, flags=re.IGNORECASE):
        for p in padroes:
            m = re.search(p, texto, flags)
            if m:
                return m.group(1).strip()
        return ""

    def proximo_paragrafo(keyword, texto):
        m = re.search(keyword, texto, re.IGNORECASE)
        if not m:
            return ""
        trecho = texto[m.end():].strip()
        linhas = [l.strip() for l in trecho.split("\n") if l.strip()]
        resultado = []
        for l in linhas:
            if re.match(r'^[A-Z0-9\s]{4,}$', l) and len(l) < 60:
                break
            resultado.append(l)
            if len(resultado) >= 3:
                break
        return " ".join(resultado)

    marca = buscar([r'Nome\s+comercial[:\s]+([^\n]+)', r'NOME\s+COMERCIAL[:\s]+([^\n]+)'], texto)
    titular = buscar([r'Titular[:\s]+([^\n]+)', r'TITULAR[:\s]+([^\n]+)',
                      r'Empresa\s+registrante[:\s]+([^\n]+)', r'Fabricante[:\s]+([^\n]+)'], texto)
    classe = buscar([r'(Isento\s+de\s+classifica[çc][aã]o\s+toxicol[oó]gica)',
                     r'(Classe\s+(?:IV|III|II|I)\b[^\n]*)',
                     r'Classifica[çc][aã]o\s+toxicol[oó]gica[:\s]+([^\n]+)'], texto)
    formulac = buscar([r'Formula[çc][aã]o[:\s]+([^\n]+)', r'FORMULA[ÇC][AÃ]O[:\s]+([^\n]+)',
                       r'Tipo\s+de\s+formula[çc][aã]o[:\s]+([^\n]+)'], texto)
    via_reg = "MAPA" if re.search(r'MAPA|Minist[eé]rio\s+da\s+Agricultura', texto) else ""
    shelf = buscar([r'[Vv]alidade[:\s]+(\d+[^\n,;.]{0,30})',
                    r'[Pp]razo\s+de\s+validade[:\s]+(\d+[^\n,;.]{0,30})',
                    r'[Ss]helf\s+[Ll]ife[:\s]+([^\n,;.]{1,30})'], texto)
    armaz = proximo_paragrafo(r'Armazenamento|Conserva[çc][aã]o|ARMAZENAMENTO', texto)
    info_tec = proximo_paragrafo(
        r'Descri[çc][aã]o\s+t[eé]cnica|Informa[çc][oõ]es\s+t[eé]cnicas|DESCRI[ÇC][AÃ]O', texto)
    ingrediente = buscar([r'Ingrediente[s]?\s+ativo[s]?[:\s]+([^\n]+)',
                          r'INGREDIENTE[S]?\s+ATIVO[S]?[:\s]+([^\n]+)',
                          r'Agente[s]?\s+biol[oó]gico[s]?[:\s]+([^\n]+)',
                          r'Microorganismo[s]?[:\s]+([^\n]+)',
                          r'Microrganismo[s]?[:\s]+([^\n]+)'], texto)
    cepa = buscar([r'[Cc][Ee][Pp][Aa][:\s]+([^\n]+)', r'[Ll]inhagem[:\s]+([^\n]+)',
                   r'[Ss]train[:\s]+([^\n]+)'], texto)
    unidade, conc_min = "", ""
    for pat_uni, uni_label in [(r'End\.\s*vi[aá]veis/ml', "End. viáveis/ml"),
                                (r'End\.\s*vi[aá]veis/L', "End. viáveis/L"),
                                (r'UFC/ml', "UFC/ml"), (r'UFC/L', "UFC/L")]:
        m = re.search(
            r'([\d.,]+\s*[xX×]\s*10\s*[\^]?\s*\d+[^\n,;]{0,10})\s*' + pat_uni, texto)
        if m:
            conc_min = m.group(1).strip() + " " + uni_label
            unidade = uni_label
            break
    micros = []
    if ingrediente:
        micros.append(dict(ingrediente=ingrediente, cepa=cepa,
                           conc_min=conc_min, unidade=unidade))
    estado_fisico = buscar([r'[Ee]stado\s+f[ií]sico[:\s]+([^\n]+)',
                             r'[Aa]specto\s+f[ií]sico[:\s]+([^\n]+)'], texto)
    cor = buscar([r'\bCor[:\s]+([^\n]+)', r'[Cc]olor[:\s]+([^\n]+)'], texto)
    aspecto = buscar([r'[Aa]specto[:\s]+([^\n]+)'], texto)
    gram = ""
    if re.search(r'[Gg]ram[- ]positiv', texto):
        gram = "Gram-positivo"
    elif re.search(r'[Gg]ram[- ]negativ', texto):
        gram = "Gram-negativo"
    morfologia = proximo_paragrafo(
        r'Morfologia|Caracter[ií]sticas\s+morfol[oó]gicas|MORFOLOGIA', texto)
    qualidade = []
    if ingrediente:
        qualidade.append(dict(ingrediente=ingrediente, estado_fisico=estado_fisico,
                              cor=cor, aspecto=aspecto, gram=gram,
                              crescimento="", meio="", morfologia=morfologia))
    return dict(marca=marca, titular=titular, info_tec=info_tec, micros=micros,
                classe=classe, formulac=formulac, via_reg=via_reg, shelf=shelf,
                armaz=armaz, rec_rows=[], qualidade=qualidade)


def extrair_qualidade_ft(texto: str) -> list:
    """Extrai dados de qualidade (Tópico 7) de uma Ficha Técnica de Qualidade.
    Retorna lista de dicts no formato qualidade[] — mesmo pipeline Mistral → Groq."""
    import re, json

    _system_ft = """Você é um especialista em fichas técnicas de qualidade de produtos biológicos agrícolas brasileiros.

REGRA ABSOLUTA: Extraia SOMENTE dados que estejam EXPLICITAMENTE no texto. NUNCA invente, NUNCA infira.
Se um campo não aparecer no texto, retorne string vazia "".

MAPEAMENTOS IMPORTANTES:
- "Estirpe" → extraia nome científico (ingrediente) e código da cepa separadamente
- "Coloração de Gram: Positiva" → gram: "Gram-positivo" | "Negativa" → "Gram-negativo"
- "Tipo de crescimento" → crescimento (ex: "Aeróbica")
- "Meio sólido de crescimento" ou "Meio de cultura" → meio
- "Características morfológicas da colônia" → morfologia (copie o texto exato)
- Extraia cor e aspecto da descrição morfológica quando explícitos (ex: "esbranquiçada" → cor, "opaco" → aspecto)
- Se o documento tiver múltiplos microrganismos em colunas paralelas, crie um objeto para cada um"""

    _prompt_ft = f"""Analise a Ficha Técnica de Qualidade abaixo e extraia os dados microbiológicos.
Retorne SOMENTE o JSON válido, sem markdown, sem explicações.

TEXTO DA FICHA TÉCNICA:
{texto[:15000]}

JSON esperado (array com um objeto por microrganismo):
[
  {{
    "ingrediente": "nome científico do microrganismo (ex: Bacillus velezensis)",
    "cepa": "código da cepa (ex: CNPSo 3602)",
    "estado_fisico": "estado físico do produto, string vazia se não houver",
    "cor": "cor do produto ou da colônia, string vazia se não houver",
    "aspecto": "aspecto visual, string vazia se não houver",
    "gram": "Gram-positivo | Gram-negativo | string vazia",
    "crescimento": "tipo de crescimento (ex: Aeróbica), string vazia se não houver",
    "meio": "meio de cultivo (ex: TSA), string vazia se não houver",
    "morfologia": "características morfológicas da colônia copiadas EXATAMENTE do texto"
  }}
]"""

    def _chamar_mistral_ft(mk):
        import requests as _req
        _r = _req.post(
            "https://api.mistral.ai/v1/chat/completions",
            headers={"Authorization": f"Bearer {mk}", "Content-Type": "application/json"},
            json={"model": "mistral-small-latest",
                  "messages": [{"role": "system", "content": _system_ft},
                                {"role": "user",   "content": _prompt_ft}],
                  "temperature": 0, "max_tokens": 4000},
            timeout=60,
        )
        _r.raise_for_status()
        return _r.json()["choices"][0]["message"]["content"].strip()

    resp = None

    # 1. Mistral (principal) — até 3 tentativas (retry em erro de API ou JSON inválido)
    try:
        mk = ""
        try:
            mk = st.secrets.get("MISTRAL_API_KEY", "")
        except Exception:
            pass
        if mk:
            import time as _time
            import json as _jft
            for _tentativa in range(3):
                try:
                    _rt = _chamar_mistral_ft(mk)
                    _rt2 = re.sub(r'^```(?:json)?\n?', '', _rt)
                    _rt2 = re.sub(r'\n?```$', '', _rt2)
                    _jft.loads(_rt2)  # valida JSON; relança exceção se inválido
                    resp = _rt
                    break
                except Exception:
                    if _tentativa < 2:
                        _time.sleep(5)
    except Exception:
        pass

    # 2. Groq fallback
    if resp is None:
        try:
            gk = ""
            try:
                gk = st.secrets.get("GROQ_API_KEY", "")
            except Exception:
                pass
            if gk:
                from groq import Groq
                gc = Groq(api_key=gk)
                r2 = gc.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "system", "content": _system_ft},
                               {"role": "user",   "content": _prompt_ft}],
                    temperature=0, max_tokens=2000,
                )
                resp = r2.choices[0].message.content.strip()
        except Exception:
            pass

    if resp is None:
        return []

    try:
        resp = re.sub(r'^```(?:json)?\n?', '', resp)
        resp = re.sub(r'\n?```$', '', resp)
        resultado = json.loads(resp)
        if isinstance(resultado, list):
            return resultado
        if isinstance(resultado, dict):
            return [resultado]
    except Exception:
        pass
    return []


def extrair_dados_bula(texto: str, uploaded_file=None) -> dict:
    import re, json

    # ── System prompt compartilhado ───────────────────────────────────────────
    _system_bula = """Você é um especialista em regulatória de produtos biológicos agrícolas no Brasil (MAPA).

REGRA MÁXIMA — SEM EXCEÇÕES:
Extraia SOMENTE dados que estejam PALAVRA POR PALAVRA no texto da bula.
Se uma informação não aparece explicitamente no texto, retorne string vazia "".
NUNCA complete, NUNCA infira, NUNCA invente dados. Em caso de dúvida, deixe vazio.

CONHECIMENTO FIXO — FORMULAÇÕES:
SC = Suspensão Concentrada (líquido), WP = Pó Molhável, WG = Granulado Dispersível,
EC = Concentrado Emulsionável, SL = Concentrado Solúvel, OD = Dispersão em Óleo,
SE = Suspoemulsão, FS = Suspensão Concentrada para Tratamento de Sementes.
SC é SEMPRE líquido/suspensão, NUNCA pó.

CONCENTRAÇÕES:
Endósporos viáveis = abreviar como End. UFC = Unidades Formadoras de Colônias.
Unidades válidas: End/ml, End/L, UFC/ml, UFC/L.
Concentração mínima: copie o valor EXATO com notação científica como aparece na bula.

CLASSE DE USO DO DEFENSIVO:
Extraia a classe de uso agronômico: Fungicida, Inseticida, Nematicida, Herbicida,
Acaricida, Bactericida, ou combinações (ex: Fungicida e Nematicida).
NÃO confundir com classificação toxicológica (Classe I, II, III, IV) — ignore essa."""

    prompt = f"""Analise o texto abaixo — ele pode conter uma BULA, uma FICHA TÉCNICA DE QUALIDADE, ou ambos.
Extraia todas as informações disponíveis no formato JSON, independente de qual documento contém cada dado.

REGRAS OBRIGATÓRIAS:
- Copie os dados EXATAMENTE como estão no texto. NÃO invente, NÃO complete, NÃO infira.
- Campo vazio no texto = string vazia "" no JSON. Nunca preencha com suposições.
- Retorne SOMENTE o JSON válido, sem markdown, sem texto adicional.

REGRA CRÍTICA PARA ALVOS BIOLÓGICOS:
Cada praga ou doença deve ser um objeto SEPARADO no array rec_rows.
ERRADO: {{"alvos": "Sclerotinia sclerotiorum, Botrytis cinerea, Fusarium sp."}}
CORRETO: três objetos separados, um para cada praga, com os mesmos valores de cultura/dose/volume/numero.
Se houver 5 alvos para a mesma cultura, crie 5 objetos. NUNCA junte alvos com vírgula ou "e".

TEXTO DO(S) DOCUMENTO(S):
{texto[:20000]}

JSON esperado:
{{
  "marca": "nome comercial EXATO como aparece no texto",
  "titular": "empresa registrante EXATA como aparece no texto",
  "n_reg_1": "SOMENTE o número/código após 'Registrado no Ministério da Agricultura e Pecuária - MAPA sob nº'. Retorne APENAS os dígitos ou código alfanumérico, SEM nenhuma descrição adicional.",
  "info_tec": "descrição técnica do produto (1 a 3 frases copiadas do texto). NÃO inclua nome do microrganismo, cepa ou concentração — esses dados ficam em campos separados. Foque em modo de ação, benefícios e uso agronômico.",
  "classe": "classe de uso agronômico EXATA (ex: Fungicida, Inseticida, Nematicida) — NÃO é classificação toxicológica",
  "formulac": "tipo de formulação por EXTENSO como no texto (ex: Suspensão Concentrada - SC)",
  "armaz": "condições de armazenamento EXATAS como no texto",
  "micros": [
    {{
      "ingrediente": "nome científico EXATO do microrganismo como no texto",
      "cepa": "código da cepa EXATO como no texto, string vazia se não houver",
      "conc_min": "APENAS o número em notação científica (ex: '1,5 x 10^11'). NÃO inclua unidade como endósporos, UFC, /L ou /ml — a unidade vai no campo 'unidade'",
      "unidade": "OBRIGATÓRIO: End/ml | End/L | UFC/ml | UFC/L — escolha conforme o texto",
      "g_l": "valor em g/L EXATO como no texto, string vazia se não aparecer",
      "m_v": "valor em m/V (%) EXATO como no texto, string vazia se não aparecer"
    }}
  ],
  "rec_rows": [
    {{
      "cultura": "cultura alvo EXATA como no texto",
      "alvos": "SOMENTE UMA praga ou doença por objeto. REGRAS CRÍTICAS: (1) Percorra TODA a tabela até o fim — não pare antes. (2) Cada alvo aparece EXATAMENTE UMA VEZ — NUNCA repita o mesmo alvo. (3) Se a tabela tem 30 alvos, o array deve ter 30 objetos. (4) Antes de retornar, verifique se há duplicatas e remova-as.",
      "dose": "Valor da coluna 'Dose', 'Dose do Produto', 'Dose (mL/ha)', 'Dose (g/ha)' ou similar na tabela de recomendações. É a quantidade do PRODUTO BIOLÓGICO a aplicar por unidade de área ou semente (ex: '100 mL/ha', '200 g/ha', '50 mL/100kg sementes'). Se for um intervalo (ex: '5 a 60 mL/100kg'), copie o intervalo COMPLETO — NUNCA escolha apenas um valor. ATENÇÃO: NÃO é o volume de calda/água.",
      "volume": "Valor da coluna 'Volume de Calda', 'Vol. de Calda', 'Volume de Aplicação', 'Volume (L/ha)' ou similar na tabela de recomendações. É a quantidade de ÁGUA ou CALDA a utilizar (ex: '200-400 L/ha', '300 mL/100L de água'). Se for um intervalo, copie o intervalo COMPLETO. Se a unidade estiver no cabeçalho da coluna, inclua-a junto ao número. ATENÇÃO: NÃO é a dose do produto.",
      "numero": "número de aplicações, época e intervalo EXATOS como no texto"
    }}
  ],
  "qualidade": [
    {{
      "ingrediente": "nome científico do microrganismo — buscar em: 'Estirpe', 'Ingrediente Ativo', nome após 'Bacillus', 'Trichoderma', etc.",
      "estado_fisico": "estado físico — buscar por: 'Estado Físico', 'Aspecto físico'. String vazia se não houver.",
      "cor": "cor do produto ou colônia — buscar por: 'Cor', ou extrair de 'Características morfológicas' (ex: 'esbranquiçada', 'amarelada'). String vazia se não houver.",
      "aspecto": "aspecto visual — buscar por: 'Aspecto', ou extrair de 'Características morfológicas' (ex: 'opaco', 'brilhante'). String vazia se não houver.",
      "gram": "buscar em 'Coloração de Gram': se 'Positiva' → 'Gram-positivo'; se 'Negativa' → 'Gram-negativo'; string vazia se não houver.",
      "crescimento": "buscar em 'Tipo de crescimento' (ex: 'Aeróbica', 'Anaeróbica', 'Facultativa'). String vazia se não houver.",
      "meio": "buscar em 'Meio sólido de crescimento', 'Meio de cultura', 'Meio de cultivo'. String vazia se não houver.",
      "morfologia": "buscar em 'Características morfológicas da colônia' — copie o texto EXATO. String vazia se não houver."
    }}
  ]
}}"""

    _UNIDADE_MAP = {
        "endósporos viáveis/ml": "End/ml", "endosporos viáveis/ml": "End/ml",
        "endósporos viáveis/l": "End/L",   "endosporos viáveis/l": "End/L",
        "end. viáveis/ml": "End/ml",       "end. viáveis/l": "End/L",
        "end. viaveis/ml": "End/ml",       "end. viaveis/l": "End/L",
        "end/ml": "End/ml",                "end/l": "End/L",
        "ufc/ml": "UFC/ml",                "ufc/l": "UFC/L",
        "unidades formadoras de colônias/ml": "UFC/ml",
        "unidades formadoras de colônias/l": "UFC/L",
        "unidades formadoras de colônia/ml": "UFC/ml",
        "unidades formadoras de colônia/l": "UFC/L",
    }

    def _normalizar_dados(dados):
        for campo in ["marca", "titular", "info_tec", "classe", "formulac", "armaz", "n_reg_1"]:
            dados.setdefault(campo, "")
        dados["shelf"] = ""
        dados.setdefault("micros", [])
        dados.setdefault("rec_rows", [])
        dados.setdefault("qualidade", [])
        for m in dados.get("micros", []):
            uni_raw = m.get("unidade", "").strip()
            m["unidade"] = _UNIDADE_MAP.get(uni_raw.lower(), uni_raw)
        # Deduplicar alvos — remove entradas com o mesmo alvo (case-insensitive)
        vistos = set()
        rec_unique = []
        for row in dados.get("rec_rows", []):
            chave = row.get("alvos", "").strip().lower()
            if chave and chave in vistos:
                continue  # duplicata — ignorar
            if chave:
                vistos.add(chave)
            rec_unique.append(row)
        dados["rec_rows"] = rec_unique
        return dados

    def _complementar(dados):
        dados["via_reg"] = "DEFENSIVO" if dados["rec_rows"] else ""
        if not dados["rec_rows"] and uploaded_file is not None:
            dados["rec_rows"] = _extrair_tabelas_recomendacoes(uploaded_file)
            if dados["rec_rows"]:
                dados["via_reg"] = "DEFENSIVO"
        return dados

    # ── Extração principal: Mistral ───────────────────────────────────────────
    mistral_key = ""
    try:
        mistral_key = st.secrets.get("MISTRAL_API_KEY", "")
    except Exception:
        pass

    def _chamar_mistral(mk, msgs):
        import requests as _req
        _r = _req.post(
            "https://api.mistral.ai/v1/chat/completions",
            headers={"Authorization": f"Bearer {mk}", "Content-Type": "application/json"},
            json={"model": "mistral-small-latest", "messages": msgs,
                  "temperature": 0, "max_tokens": 6000},
            timeout=60,
        )
        _r.raise_for_status()
        return _r.json()["choices"][0]["message"]["content"].strip()

    if mistral_key:
        try:
            import time as _time_m
            dados = None
            for _tent in range(3):
                try:
                    resposta_m = _chamar_mistral(mistral_key, [
                        {"role": "system", "content": _system_bula},
                        {"role": "user",   "content": prompt}
                    ])
                    _rm = re.sub(r'^```(?:json)?\n?', '', resposta_m)
                    _rm = re.sub(r'\n?```$', '', _rm)
                    dados = json.loads(_rm)
                    break
                except Exception:
                    if _tent < 2:
                        _time_m.sleep(5)
            if dados is None:
                raise Exception("Mistral falhou após 3 tentativas")
            dados = _normalizar_dados(dados)

            # Revisão com Mistral
            try:
                prompt_rev_m = f"""Você é um revisor rigoroso de extração de dados de documentos agrícolas.
REGRA MÁXIMA: Só corrija se a informação estiver EXPLICITAMENTE no texto. NUNCA invente.
1. Corrija erros claros; 2. Complete campos vazios só se estiver no texto;
3. Separe alvos agrupados (um objeto por alvo); 4. Verifique g/L e m/V;
5. Verifique classe de uso (Fungicida/Inseticida/etc); 6. Verifique qualidade (Gram, crescimento, meio, morfologia)
Retorne o JSON COMPLETO no mesmo formato, sem markdown.

DADOS EXTRAÍDOS:
{json.dumps(dados, ensure_ascii=False, indent=2)}

TEXTO ORIGINAL:
{texto[:15000]}"""
                resp_rev = _chamar_mistral(mistral_key, [
                    {"role": "system", "content": _system_bula},
                    {"role": "user",   "content": prompt_rev_m}
                ])
                resp_rev = re.sub(r'^```(?:json)?\n?', '', resp_rev)
                resp_rev = re.sub(r'\n?```$', '', resp_rev)
                dados_rev = json.loads(resp_rev)
                for campo in ["marca","titular","info_tec","classe","formulac","armaz"]:
                    if dados_rev.get(campo): dados[campo] = dados_rev[campo]
                if dados_rev.get("micros") and len(dados_rev["micros"]) >= len(dados.get("micros",[])):
                    dados["micros"] = dados_rev["micros"]
                if dados_rev.get("rec_rows") and len(dados_rev["rec_rows"]) >= len(dados.get("rec_rows",[])):
                    dados["rec_rows"] = dados_rev["rec_rows"]
                if dados_rev.get("qualidade"): dados["qualidade"] = dados_rev["qualidade"]
                st.session_state["_debug_revisao"] = True
            except Exception:
                st.session_state["_debug_revisao"] = False

            for m in dados.get("micros", []):
                uni_raw = m.get("unidade", "").strip()
                m["unidade"] = _UNIDADE_MAP.get(uni_raw.lower(), uni_raw)
            dados = _complementar(dados)
            st.session_state["_debug_mistral"] = True
            return dados
        except Exception as em:
            st.warning(f"Mistral indisponível, tentando Groq. Detalhe: {em}")

    # ── Fallback: Groq ────────────────────────────────────────────────────────
    groq_key = ""
    try:
        groq_key = st.secrets.get("GROQ_API_KEY", "")
    except Exception:
        pass

    if groq_key:
        try:
            from groq import Groq
            client = Groq(api_key=groq_key)

            # 1ª chamada: extração com modelo maior
            response = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": _system_bula},
                    {"role": "user", "content": prompt}
                ],
                temperature=0,
                max_tokens=3000,
            )
            resposta = response.choices[0].message.content.strip()
            resposta = re.sub(r'^```(?:json)?\n?', '', resposta)
            resposta = re.sub(r'\n?```$', '', resposta)
            dados = json.loads(resposta)
            dados = _normalizar_dados(dados)

            # 2ª chamada: revisão — verifica campos vazios e erros óbvios
            try:
                prompt_rev = f"""Você é um revisor rigoroso de extração de dados de documentos agrícolas (bula e/ou ficha técnica).

REGRA MÁXIMA: Só corrija ou preencha um campo se a informação estiver EXPLICITAMENTE no texto abaixo.
Se não tiver certeza absoluta de que o dado está no texto, mantenha o campo como está ou deixe vazio.
NUNCA invente, NUNCA infira, NUNCA complete com conhecimento próprio.

Sua tarefa:
1. Compare os dados extraídos com o texto original
2. Corrija apenas erros claros (ex: número trocado, nome grafado errado)
3. Preencha campos vazios SOMENTE se o dado aparecer literalmente no texto
4. CRÍTICO — Verifique os campos de qualidade (array "qualidade"). Se estiverem vazios, procure no texto por:
   - "Estirpe" → ingrediente (nome científico) e cepa (código após o traço "-")
   - "Coloração de Gram: Positiva" → gram: "Gram-positivo" | "Negativa" → "Gram-negativo"
   - "Tipo de crescimento" → crescimento
   - "Meio sólido de crescimento" ou "Meio de cultura" → meio
   - "Características morfológicas da colônia" → morfologia (copie o texto exato)
   - Extraia cor e aspecto da descrição morfológica (ex: "esbranquiçada" → cor, "opaco" → aspecto)
5. CRÍTICO — Separe alvos agrupados: se um campo "alvos" contém múltiplas pragas separadas por vírgula ou "e",
   crie um objeto separado para cada praga, repetindo os valores de cultura/dose/volume/numero.
   ERRADO: {{"alvos": "Praga A, Praga B"}} — CORRETO: dois objetos, um com "Praga A" e outro com "Praga B"
5. Verifique se g/L e m/V estão no texto — se sim, preencha; se não, deixe vazio
6. Verifique se a classe de uso (Fungicida, Inseticida, etc.) está correta — NÃO é Classe I/II/III/IV
7. CRÍTICO — Verifique se dose e volume NÃO foram invertidos. Localize a tabela de recomendações no texto original:
   - O valor na coluna "Dose" / "Dose do Produto" deve estar em rec_rows[*].dose
   - O valor na coluna "Volume de Calda" / "Vol. de Calda" deve estar em rec_rows[*].volume
   - dose = quantidade do PRODUTO BIOLÓGICO (ex: mL/ha, g/ha, mL/kg sementes)
   - volume = quantidade de ÁGUA/CALDA (ex: L/ha, mL/100L de água)
   - REGRA INFALÍVEL: se o valor contiver a palavra "água", "calda" ou "diluente" → pertence a rec_rows[*].volume, NUNCA a dose
   Se estiverem trocados, corrija comparando com o texto original.
8. Retorne o JSON COMPLETO no mesmo formato, sem markdown, sem explicações

DADOS EXTRAÍDOS:
{json.dumps(dados, ensure_ascii=False, indent=2)}

TEXTO ORIGINAL:
{texto[:15000]}"""

                # Revisão: Groq 70b → Groq 8b → Mistral (à prova de falhas)
                _resp2_text = None

                # 1. Groq 70b
                try:
                    _r2 = client.chat.completions.create(
                        model="llama-3.3-70b-versatile",
                        messages=[{"role": "system", "content": _system_bula},
                                  {"role": "user",   "content": prompt_rev}],
                        temperature=0, max_tokens=3000,
                    )
                    _resp2_text = _r2.choices[0].message.content.strip()
                except Exception:
                    pass

                # 2. Groq 8b fallback
                if _resp2_text is None:
                    try:
                        _r2b = client.chat.completions.create(
                            model="llama-3.1-8b-instant",
                            messages=[{"role": "system", "content": _system_bula},
                                      {"role": "user",   "content": prompt_rev}],
                            temperature=0, max_tokens=3000,
                        )
                        _resp2_text = _r2b.choices[0].message.content.strip()
                    except Exception:
                        pass

                # 3. Mistral fallback
                if _resp2_text is None:
                    try:
                        _mk = ""
                        try:
                            _mk = st.secrets.get("MISTRAL_API_KEY", "")
                        except Exception:
                            pass
                        if _mk:
                            import requests as _req
                            _rm = _req.post(
                                "https://api.mistral.ai/v1/chat/completions",
                                headers={"Authorization": f"Bearer {_mk}",
                                         "Content-Type": "application/json"},
                                json={"model": "mistral-small-latest",
                                      "messages": [{"role": "system", "content": _system_bula},
                                                   {"role": "user",   "content": prompt_rev}],
                                      "temperature": 0, "max_tokens": 3000},
                                timeout=60,
                            )
                            _rm.raise_for_status()
                            _resp2_text = _rm.json()["choices"][0]["message"]["content"].strip()
                    except Exception:
                        pass

                # Aplicar revisão se obteve resposta
                if _resp2_text:
                    try:
                        _resp2_text = re.sub(r'^```(?:json)?\n?', '', _resp2_text)
                        _resp2_text = re.sub(r'\n?```$', '', _resp2_text)
                        dados_rev = json.loads(_resp2_text)
                        for campo in ["marca", "titular", "info_tec", "classe", "formulac", "armaz"]:
                            if dados_rev.get(campo):
                                dados[campo] = dados_rev[campo]
                        if dados_rev.get("micros") and len(dados_rev["micros"]) >= len(dados.get("micros", [])):
                            dados["micros"] = dados_rev["micros"]
                        if dados_rev.get("rec_rows") and len(dados_rev["rec_rows"]) >= len(dados.get("rec_rows", [])):
                            dados["rec_rows"] = dados_rev["rec_rows"]
                        if dados_rev.get("qualidade"):
                            dados["qualidade"] = dados_rev["qualidade"]
                        st.session_state["_debug_revisao"] = True
                    except Exception:
                        st.session_state["_debug_revisao"] = False
                else:
                    st.session_state["_debug_revisao"] = False

            except Exception:
                st.session_state["_debug_revisao"] = False

            for m in dados.get("micros", []):
                uni_raw = m.get("unidade", "").strip()
                m["unidade"] = _UNIDADE_MAP.get(uni_raw.lower(), uni_raw)

            dados = _complementar(dados)
            return dados

        except Exception as e:
            if "rate_limit" not in str(e).lower() and "429" not in str(e):
                st.warning(f"Groq indisponível. Detalhe: {e}")
            # Se for rate limit, tenta Gemini antes de ir para regex

    # ── Fallback final: extração por regex ────────────────────────────────────
    dados = _extrair_dados_bula_regex(texto)
    if uploaded_file is not None:
        dados["rec_rows"] = _extrair_tabelas_recomendacoes(uploaded_file)
    return dados


def popular_session_state(dados: dict):
    for campo in ["marca", "titular", "info_tec", "classe", "formulac",
                  "via_reg", "shelf", "armaz", "n_reg_1"]:
        if dados.get(campo):
            st.session_state[campo] = dados[campo]

    micros = dados.get("micros", [])
    if micros:
        st.session_state["qtd_micro"] = min(len(micros), 6)
        for i, m in enumerate(micros[:6]):
            st.session_state[f"ing_{i}"]  = m.get("ingrediente", "")
            st.session_state[f"cepa_{i}"] = m.get("cepa", "")
            st.session_state[f"cmin_{i}"] = m.get("conc_min", "")
            st.session_state[f"gl_{i}"]   = m.get("g_l", "")
            st.session_state[f"mv_{i}"]   = m.get("m_v", "")
            st.session_state[f"uni_{i}"]  = m.get("unidade", "End/ml")

    rec_rows = dados.get("rec_rows", [])
    if rec_rows:
        st.session_state["n_rec"] = len(rec_rows)
        for i, r in enumerate(rec_rows):
            # cult_ não é sobrescrito — sempre "Em todas as Culturas"
            st.session_state[f"alv_{i}"]  = str(r.get("alvos", ""))
            st.session_state[f"dose_{i}"] = str(r.get("dose", ""))
            st.session_state[f"vol_{i}"]  = str(r.get("volume", ""))
            st.session_state[f"num_{i}"]  = str(r.get("numero", ""))

    for i, q in enumerate(dados.get("qualidade", [])):
        for key, campo in [("q_ing", "ingrediente"), ("q_est", "estado_fisico"),
                           ("q_cor", "cor"), ("q_asp", "aspecto"),
                           ("q_gram", "gram"), ("q_cresc", "crescimento"),
                           ("q_meio", "meio"), ("q_morf", "morfologia")]:
            st.session_state[f"{key}_{i}"] = q.get(campo, "")

    # Sinaliza que precisa de um segundo rerun para widgets novos lerem o session_state
    st.session_state["_popular_rerun"] = True
    st.rerun()


# ── BACKEND — leitura de dimensionamento de embalagens ────────────────────────
def carregar_backend(embalagem: str) -> dict:
    try:
        backend_path = os.path.join(os.path.dirname(TEMPLATE_PATH), "BACKEND_extratos.xlsx")
        if not os.path.exists(backend_path):
            return {}
        wb_bk = openpyxl.load_workbook(backend_path, data_only=True)
        ws_bk = wb_bk["BACKEND"]
        for row in ws_bk.iter_rows(min_row=3, max_row=10):
            if str(row[0].value or "").strip() == embalagem.strip():
                vals = [cell.value for cell in row]
                r = {}
                for i in range(6):
                    r[f"t21_dim{i+1}"] = str(vals[1+i*2] or "") if len(vals) > 1+i*2 else ""
                    r[f"t21_med{i+1}"] = str(vals[2+i*2] or "") if len(vals) > 2+i*2 else ""
                r["cx_int_nome"] = str(vals[13] or "") if len(vals) > 13 else ""
                for i in range(6):
                    r[f"t23_dim{i+1}"] = str(vals[14+i*2] or "") if len(vals) > 14+i*2 else ""
                    r[f"t23_med{i+1}"] = str(vals[15+i*2] or "") if len(vals) > 15+i*2 else ""
                r["cx_mst_qtd"] = str(vals[26] or "") if len(vals) > 26 else ""
                for i in range(6):
                    r[f"t25_dim{i+1}"] = str(vals[27+i*2] or "") if len(vals) > 27+i*2 else ""
                    r[f"t25_med{i+1}"] = str(vals[28+i*2] or "") if len(vals) > 28+i*2 else ""
                r["t26_tipo"] = str(vals[39] or "") if len(vals) > 39 else ""
                return r
        return {}
    except Exception:
        return {}


# ── Helpers ───────────────────────────────────────────────────────────────────
def sec(titulo):
    st.markdown(f'<div class="sec-header">{titulo}</div>', unsafe_allow_html=True)

def step_info(numero, descricao, proxima=None):
    prox_txt = f" &nbsp;|&nbsp; Próxima: <b>{proxima}</b> →" if proxima else ""
    st.markdown(
        f'<div class="step-guide">Passo <b>{numero} de 11</b> — {descricao}{prox_txt}</div>',
        unsafe_allow_html=True
    )

def sn(val: bool) -> str:
    return "SIM" if val else "NÃO"

def escrever(ws, cel, val, wrap=False):
    try:
        ws[cel] = val
        ws[cel].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=wrap
        )
        ws[cel].font = Font(color="595959")
        if val == "SIM":
            ws[cel].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                                       fill_type="solid")
        elif val == "NÃO":
            ws[cel].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                       fill_type="solid")
    except AttributeError:
        pass

def inserir_imagem_centralizada(ws, img_bytes_or_path, anchor_cel, max_w=None, max_h=None):
    """Insere imagem redimensionada proporcionalmente e centralizada na célula anchor.

    Detecta automaticamente intervalos mesclados e soma todas as colunas/linhas
    do intervalo para calcular as dimensões reais da área de destino.
    max_w e max_h são mantidos apenas para compatibilidade — não são mais usados.
    """
    try:
        import re as _re2
        from PIL import Image as PILImg
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils import column_index_from_string, get_column_letter
        from openpyxl.utils.units import pixels_to_EMU

        if isinstance(img_bytes_or_path, (bytes, bytearray)):
            pil    = PILImg.open(io.BytesIO(img_bytes_or_path))
            xl_img = XLImage(io.BytesIO(img_bytes_or_path))
        else:
            pil    = PILImg.open(img_bytes_or_path)
            xl_img = XLImage(img_bytes_or_path)

        w, h = pil.size

        # Parsear coordenada da célula âncora (ex: "C10")
        m = _re2.match(r'([A-Z]+)(\d+)', anchor_cel)
        col_letter = m.group(1)
        row_num    = int(m.group(2))
        col_1based = column_index_from_string(col_letter)

        # Detectar intervalo mesclado que contém a célula
        min_col, max_col = col_1based, col_1based
        min_row, max_row = row_num, row_num
        for merge in ws.merged_cells.ranges:
            if (merge.min_col <= col_1based <= merge.max_col and
                    merge.min_row <= row_num <= merge.max_row):
                min_col, max_col = merge.min_col, merge.max_col
                min_row, max_row = merge.min_row, merge.max_row
                break

        # Dimensões reais em pixels, somando todas as colunas/linhas do intervalo.
        # Lê as dimensões sem criar entradas novas no holder (evita sobrescrever
        # larguras implícitas do template ao salvar, o que quebra o layout A4).
        _existing_cols = set(ws.column_dimensions.keys())
        _existing_rows = set(ws.row_dimensions.keys())

        cell_w_px = sum(
            int((ws.column_dimensions[get_column_letter(c)].width or 8) * 7)
            for c in range(min_col, max_col + 1)
        )
        cell_h_px = sum(
            (ws.row_dimensions[r].height or 15) * 96 / 72
            for r in range(min_row, max_row + 1)
        )

        # Remover entradas criadas artificialmente pelo acesso acima
        for _c in list(ws.column_dimensions.keys()):
            if _c not in _existing_cols:
                del ws.column_dimensions[_c]
        for _r in list(ws.row_dimensions.keys()):
            if _r not in _existing_rows:
                del ws.row_dimensions[_r]

        # Margem de 10% para garantir que a imagem não vaze para fora da célula
        eff_w = cell_w_px * 0.90
        eff_h = cell_h_px * 0.90
        # Nunca ampliar (ratio máximo = 1.0), só reduzir se necessário
        ratio = min(eff_w / w, eff_h / h, 1.0)
        img_w = int(w * ratio)
        img_h = int(h * ratio)

        # Offset para centralizar dentro do intervalo
        x_off = max(0, (cell_w_px - img_w) / 2)
        y_off = max(0, (cell_h_px - img_h) / 2)

        marker = AnchorMarker(
            col=col_1based - 1, colOff=pixels_to_EMU(x_off),  # AnchorMarker é 0-based
            row=min_row - 1,    rowOff=pixels_to_EMU(y_off)
        )
        size = XDRPositiveSize2D(pixels_to_EMU(img_w), pixels_to_EMU(img_h))
        xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
        # Não definir xl_img.width/height: o tamanho já está no anchor (ext)
        # e definir aqui sobrescreveria o valor em EMU com pixels brutos.
        ws.add_image(xl_img)
    except Exception:
        pass

def escrever_link(ws, cel, nome_arquivo):
    try:
        ws[cel].value = nome_arquivo
        ws[cel].font = Font(color="0563C1", underline="single")
        ws[cel].alignment = Alignment(horizontal="center", vertical="center",
                                      wrap_text=False)
    except AttributeError:
        pass

def preservar_imagens_bytes(template_bytes: bytes, output_bytes: bytes) -> bytes:
    """Restaura imagens/desenhos do template que openpyxl descartou ao salvar."""
    partes = ("xl/media/", "xl/drawings/", "xl/drawings/_rels/",
              "xl/worksheets/_rels/")
    with zipfile.ZipFile(io.BytesIO(template_bytes), "r") as t:
        tmpl = {n: t.read(n) for n in t.namelist()}
    with zipfile.ZipFile(io.BytesIO(output_bytes), "r") as o:
        out = {n: o.read(n) for n in o.namelist()}
    merged = dict(out)
    for name, data in tmpl.items():
        if name not in out and any(name.startswith(p) for p in partes):
            merged[name] = data
    result = io.BytesIO()
    with zipfile.ZipFile(result, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in merged.items():
            z.writestr(name, data)
    result.seek(0)
    return result.getvalue()

# ── Proteção contra perda de progresso ────────────────────────────────────────
components.html("""
<script>
window.addEventListener('beforeunload', function (e) {
    e.preventDefault();
    e.returnValue = 'Você tem dados não salvos. Deseja sair?';
});
</script>
""", height=0)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="biotrop-header" style="text-align:center">
  <h1 style="margin:0;padding:0;color:#2D2F33;font-size:28px;letter-spacing:2px;line-height:1.1;font-family:'Space Grotesk',sans-serif;font-weight:700">ATLAS</h1>
  <h2 style="margin:2px 0 0 0;padding:0;color:#69695D;font-size:13px;font-weight:400;line-height:1.2;font-family:'Space Grotesk',sans-serif">Agronomic, Technical & Legal Attributes Sheet</h2>
  <p style="margin:6px 0 0 0;padding:0;color:#52b788;font-size:10px;opacity:0.7;font-style:italic;line-height:1">Created by: Eduardo Rossatto</p>
</div>
""", unsafe_allow_html=True)

# ── Navegação entre abas ──────────────────────────────────────────────────────
if "next_tab" not in st.session_state:
    st.session_state.next_tab = None

# ── Upload de Bula para preenchimento automático ───────────────────────────────
with st.expander("Preencher automaticamente via PDF"):
    st.caption("Faça upload da bula e/ou da ficha técnica de qualidade. Os dois são opcionais — envie o que tiver disponível.")
    c1, c2 = st.columns(2)
    bula_pdf    = c1.file_uploader("Bula (PDF)", type=["pdf"], key="bula_upload")
    ficha_pdf   = c2.file_uploader("Ficha Técnica de Qualidade (PDF ou Excel)", type=["pdf", "xlsx"], key="ficha_upload")
    if st.button("Extrair dados", key="btn_extrair"):
        if bula_pdf is None and ficha_pdf is None:
            st.warning("Selecione ao menos um arquivo antes de extrair.")
        else:
            with st.spinner("Lendo arquivo(s) e extraindo dados..."):
                try:
                    partes = []
                    texto_ficha = ""
                    if bula_pdf is not None:
                        partes.append("=== BULA DO PRODUTO ===\n" + extrair_texto_pdf(bula_pdf))
                    if ficha_pdf is not None:
                        if ficha_pdf.name.endswith(".xlsx"):
                            texto_ficha = extrair_texto_excel(ficha_pdf)
                        else:
                            texto_ficha = extrair_texto_pdf(ficha_pdf)
                        partes.append("=== FICHA TÉCNICA DE QUALIDADE ===\n" + texto_ficha)
                    texto = "\n\n".join(partes)
                    groq_key = ""
                    try:
                        groq_key = st.secrets.get("GROQ_API_KEY", "")
                    except Exception:
                        pass
                    dados = extrair_dados_bula(texto, uploaded_file=bula_pdf)

                    # Extração dedicada de qualidade da ficha técnica
                    if ficha_pdf is not None and texto_ficha.strip():
                        qualidade_ft = extrair_qualidade_ft(texto_ficha)
                        if qualidade_ft:
                            dados["qualidade"] = qualidade_ft

                    st.session_state["_dados_extraidos"] = dados
                    st.session_state["_debug_chars"] = len(texto)
                    st.session_state["_debug_groq"] = bool(groq_key)
                except Exception as e:
                    st.error(f"Erro na extração: {e}")

    if st.session_state.get("_dados_extraidos"):
        dados = st.session_state["_dados_extraidos"]
        chars = st.session_state.get("_debug_chars", 0)
        groq_ok    = st.session_state.get("_debug_groq", False)
        revisao_ok = st.session_state.get("_debug_revisao", False)
        mistral_ok = st.session_state.get("_debug_mistral", False)
        st.success("Dados extraídos! Confira abaixo antes de preencher.")
        if mistral_ok:
            status_ia = "✓ Mistral (extração)"
        elif groq_ok and revisao_ok:
            status_ia = "✓ Groq (extração + revisão)"
        elif groq_ok:
            status_ia = "✓ Groq (extração, sem revisão)"
        else:
            status_ia = "✗ usando regex"
        st.caption(f"PDF: {chars} caracteres | IA: {status_ia}")
        with st.expander("Ver dados extraídos"):
            st.json(dados)
        if st.button("Preencher formulário com esses dados", key="btn_popular"):
            popular_session_state(dados)

# Segundo rerun automático — garante que widgets novos leiam o session_state
if st.session_state.pop("_popular_rerun", False):
    st.rerun()

# ── Tabs ──────────────────────────────────────────────────────────────────────
t1, t2, t3, t4, t5, t6, t7, t8, t9, t10 = st.tabs(ABAS)

# ── Aba 1 — Identificação ─────────────────────────────────────────────────────
with t1:
    step_info(1, "Identificação do produto, titular e informações técnicas", proxima="2. Concentração")

    sec("CABEÇALHO")
    c1, c2 = st.columns(2)
    data      = c1.text_input("Data", value=date.today().strftime("%d/%m/%Y"))
    cod_curto = c2.text_input("Código da Ficha")

    sec("1. IDENTIFICAÇÃO DO PRODUTO")
    _PPT_OPTS = ["", "8x1,5L", "4x3L", "12x1L", "4x5L", "1x20L"]
    c1, c2 = st.columns(2)
    cod_int = c1.text_input("Código do Produto")
    marca   = c2.text_input("Marca Comercial", key="marca")
    qtd_cod_pa = st.radio("Quantos Códigos P.A.?", [1, 2], horizontal=True, key="qtd_cod_pa")
    c1, c2 = st.columns(2)
    cod_pa_ppt1 = c1.selectbox("Apresentação (P.A. 1)", _PPT_OPTS, key="cod_pa_ppt1")
    cod_pa      = c2.text_input("Código P.A. 1")
    cod_pa_2 = ""
    cod_pa_3 = ""
    cod_pa_ppt2 = ""
    if qtd_cod_pa == 2:
        c1, c2 = st.columns(2)
        cod_pa_ppt2 = c1.selectbox("Apresentação (P.A. 2)", _PPT_OPTS, key="cod_pa_ppt2")
        cod_pa_2    = c2.text_input("Código P.A. 2", key="cod_pa_2")

    sec("NÚMERO DE REGISTRO")
    n_reg_1 = st.text_input("Número de Registro (MAPA)", key="n_reg_1")

    sec("4. TITULAR DE REGISTRO")
    titular = st.text_input("Titular", key="titular")

    sec("5. INFORMAÇÕES TÉCNICAS SOBRE O PRODUTO")
    info_tec = st.text_area("Descrição técnica", height=100, placeholder="opcional", key="info_tec")

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_1", use_container_width=True):
        st.session_state.next_tab = 1
        st.rerun()

# ── Aba 2 — Concentração ──────────────────────────────────────────────────────
with t2:
    step_info(2, "Concentrações e garantias", proxima="3. Classificação")

    sec("2. CONCENTRAÇÕES E GARANTIAS")
    qtd_micro = st.radio("Quantos microrganismos?", [1, 2, 3, 4, 5, 6], horizontal=True, key="qtd_micro")

    micros = []
    for i in range(qtd_micro):
        titulo = f"Microrganismo {i+1}" if qtd_micro > 1 else "Dados do microrganismo"
        sec(titulo)
        c1, c2 = st.columns(2)
        ing  = c1.text_input("Ingrediente ativo", key=f"ing_{i}")
        cepa = c2.text_input("CEPA",              key=f"cepa_{i}")
        c1, c2 = st.columns(2)
        cmin = c1.text_input("Conc. Mínima (ex: 1,5 x 10^11)", key=f"cmin_{i}")
        uni  = c2.selectbox("Unidade", UNIDADES, key=f"uni_{i}")
        c1, c2 = st.columns(2)
        g_l  = c1.text_input("g/L",  placeholder="opcional", key=f"gl_{i}")
        m_v  = c2.text_input("m/V",  placeholder="opcional", key=f"mv_{i}")
        micros.append(dict(ingrediente=ing, cepa=cepa, conc_min=cmin,
                           unidade=uni, g_l=g_l, m_v=m_v))

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_2", use_container_width=True):
        st.session_state.next_tab = 2
        st.rerun()

# ── Aba 3 — Classificação ─────────────────────────────────────────────────────
with t3:
    step_info(3, "Classificações regulatórias e fiscais", proxima="4. Recomendações")

    sec("3. CLASSIFICAÇÕES REGULATÓRIAS E FISCAIS")
    c1, c2 = st.columns(2)
    classe   = c1.text_input("Classe", key="classe")
    formulac = c2.text_input("Formulação", key="formulac")
    c1, c2   = st.columns(2)
    via_reg  = c1.text_input("Via de registro", key="via_reg")
    shelf    = c2.text_input("Shelf life", key="shelf")
    armaz = ""
    _NCM_OPTS = [
        "",
        "BIOFUNGICIDA - 3808.92.99",
        "BIONEMATICIDA - 3808.94.29",
        "BIOINSETICIDA - 3808.91.99",
        "BIOHERBICIDA - 3808.93.99",
        "INOCULANTE - 3002.49.99",
    ]
    ncm = st.selectbox("NCM", _NCM_OPTS)

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_3", use_container_width=True):
        st.session_state.next_tab = 3
        st.rerun()

# ── Aba 4 — Recomendações ─────────────────────────────────────────────────────
with t4:
    step_info(4, "Recomendações de uso por cultura", proxima="5. Qualidade")

    sec("6. RECOMENDAÇÕES DE USO")
    st.caption('Todos os campos de Cultura estão pré-preenchidos com "Todas as culturas com ocorrência do alvo biológico".')

    if "n_rec" not in st.session_state:
        st.session_state.n_rec = 1
    if st.button("+ Adicionar linha"):
        st.session_state.n_rec += 1

    rec_rows = []
    cols = st.columns([2, 2, 1, 1, 2])
    for col, lbl in zip(cols, ["Cultura", "Alvos Biológicos", "Dose",
                                "Vol. Calda", "Nº/Época/Intervalo"]):
        col.markdown(f"**{lbl}**")

    for i in range(st.session_state.n_rec):
        c1, c2, c3, c4, c5 = st.columns([2, 2, 1, 1, 2])
        # Pré-preencher cultura com "Em todas as Culturas" se ainda não foi definida
        if f"cult_{i}" not in st.session_state:
            st.session_state[f"cult_{i}"] = "Todas as culturas com ocorrência do alvo biológico"
        cultura = c1.text_input("", key=f"cult_{i}",
                                label_visibility="collapsed")
        alvos = c2.text_input("", placeholder="Alvos",  key=f"alv_{i}",
                               label_visibility="collapsed")
        dose  = c3.text_input("", placeholder="Dose",   key=f"dose_{i}",
                               label_visibility="collapsed")
        vol   = c4.text_input("", placeholder="Volume", key=f"vol_{i}",
                               label_visibility="collapsed")
        num   = c5.text_input("", placeholder="Nº/Época", key=f"num_{i}",
                               label_visibility="collapsed")
        rec_rows.append(dict(cultura=cultura, alvos=alvos,
                             dose=dose, volume=vol, numero=num))

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_4", use_container_width=True):
        st.session_state.next_tab = 4
        st.rerun()

# ── Aba 5 — Qualidade ─────────────────────────────────────────────────────────
with t5:
    step_info(5, "Qualidade, análises físico-químicas e morfologia", proxima="6. CoA / Estabilidade")

    # 7. Ficha técnica de qualidade
    qualidade = []
    for i in range(qtd_micro):
        titulo = f"7. FICHA TÉCNICA DE QUALIDADE — Microrganismo {i+1}" \
                 if qtd_micro > 1 else "7. FICHA TÉCNICA DE QUALIDADE"
        sec(titulo)
        c1, c2 = st.columns(2)
        _ing_default = micros[i]["ingrediente"] if i < len(micros) and micros[i]["ingrediente"] else ""
        if f"q_ing_{i}" not in st.session_state:
            st.session_state[f"q_ing_{i}"] = _ing_default
        ing_q = c1.text_input("Ingrediente ativo",   placeholder="opcional", key=f"q_ing_{i}")
        est_f = c2.text_input("Estado físico",        placeholder="opcional", key=f"q_est_{i}")
        c1, c2 = st.columns(2)
        cor   = c1.text_input("Cor",                 placeholder="opcional", key=f"q_cor_{i}")
        asp   = c2.text_input("Aspecto",              placeholder="opcional", key=f"q_asp_{i}")
        c1, c2 = st.columns(2)
        gram  = c1.text_input("Coloração de Gram",   placeholder="opcional", key=f"q_gram_{i}")
        cresc = c2.text_input("Tipo de crescimento", placeholder="opcional", key=f"q_cresc_{i}")
        meio  = st.text_input("Meio sólido",          placeholder="opcional", key=f"q_meio_{i}")
        morf  = st.text_area("Características morfológicas", placeholder="opcional",
                              key=f"q_morf_{i}", height=80)
        qualidade.append(dict(ingrediente=ing_q, estado_fisico=est_f, cor=cor,
                              aspecto=asp, gram=gram, crescimento=cresc,
                              meio=meio, morfologia=morf))

    # 8. Resultado das análises físico-químicas e microbiológicas
    sec("8. RESULTADO DAS ANÁLISES FÍSICO-QUÍMICAS E MICROBIOLÓGICAS")
    analises_nomes = ["Análise de Pureza", "Análise de Concentração", "PH",
                      "Densidade", "Viscosidade", "Teor de umidade/água/granulometria"]
    h1, h2, h3, h4 = st.columns([3, 1, 1, 1])
    for col, lbl in zip([h1,h2,h3,h4], ["Análise","Aplicável?","Mínimo","Máximo"]):
        col.markdown(f"**{lbl}**")
    analises = []
    unidade_conc = ""
    for nome in analises_nomes:
        if nome == "Análise de Concentração":
            c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
            _name_col, _unit_col = c1.columns([2, 1])
            _name_col.write(nome)
            unidade_conc = _unit_col.selectbox("", ["End/ml", "End/L", "UFC/ml", "UFC/L"],
                                               key="unidade_conc", label_visibility="collapsed")
            _l, _m, _r = c2.columns([1, 2, 1])
            aplic = _m.checkbox("", key=f"ap_{nome}", label_visibility="collapsed")
            vmin = c3.text_input("", key=f"mn_{nome}", label_visibility="collapsed", placeholder="mín")
            vmax = c4.text_input("", key=f"mx_{nome}", label_visibility="collapsed", placeholder="máx")
        else:
            c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
            c1.write(nome)
            _l, _m, _r = c2.columns([1, 2, 1])
            aplic = _m.checkbox("", key=f"ap_{nome}", label_visibility="collapsed")
            vmin  = c3.text_input("", key=f"mn_{nome}", label_visibility="collapsed", placeholder="mín")
            vmax  = c4.text_input("", key=f"mx_{nome}", label_visibility="collapsed", placeholder="máx")
        analises.append((aplic, vmin, vmax))

    # 10. Fotos e morfologia
    sec("10. FOTOS E MORFOLOGIA DAS COLÔNIAS E MICROSCOPIA")
    imgs_fig  = {}
    descs_fig = {}
    for row_start in (1, 4):
        c1, c2, c3 = st.columns(3)
        for i, col in enumerate([c1, c2, c3], start=row_start):
            with col:
                ct = st.container(border=True)
                with ct:
                    st.markdown(f"**Figura {i}**")
                    imgs_fig[i]  = st.file_uploader(f"Imagem {i}", type=["png","jpg","jpeg","bmp","tiff"], key=f"img_{i}")
                    descs_fig[i] = st.text_area(f"Descrição {i}", height=60, key=f"desc_{i}", placeholder="opcional")

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_5", use_container_width=True):
        st.session_state.next_tab = 5
        st.rerun()

# ── Aba 6 — CoA / Estabilidade ────────────────────────────────────────────────
with t6:
    step_info(6, "Certificado de análise e estudos de estabilidade", proxima="7. Documentos")

    sec("11. CERTIFICADO DE ANÁLISE (CoA)")
    st.caption("Defina quais análises são obrigatórias no CoA deste produto. Somente os itens selecionados serão exigidos a cada lote.")

    st.markdown("##### Verificação de concentração")
    c1, c2 = st.columns(2)
    coa_i  = c1.checkbox("Concentração de Garantia")
    coa_ii = c2.checkbox("Concentração do Lote")

    st.markdown("##### Análises físico-químicas obrigatórias no CoA")
    _coa_iii_opts = ["Análise de Pureza",
                     "PH", "Densidade", "Viscosidade",
                     "Teor de umidade/água/granulometria"]
    coa_iii_selecionadas = st.multiselect("Selecione as análises", _coa_iii_opts, key="coa_iii_sel")

    st.markdown("##### Microbiologia e segurança")
    coa_iv = st.checkbox("Contaminantes patogênicos")

    st.markdown("##### Justificativa para itens não aplicáveis")
    justif = st.text_area("", height=80, placeholder="Descreva o motivo caso alguma análise não se aplique (opcional)", label_visibility="collapsed")

    sec("11.1 ESTUDOS DE ESTABILIDADE")
    h1, h2, h3 = st.columns([2, 1, 2])
    h1.markdown("**Modalidade**"); h2.markdown("**Aplicável?**"); h3.markdown("**Tempo**")
    c1, c2, c3 = st.columns([2, 1, 2])
    c1.write("Estabilidade acelerada")
    _l, _m, _r = c2.columns([1, 2, 1])
    estab_ac_aplic = _m.checkbox("", key="eac", label_visibility="collapsed")
    estab_ac_tempo = c3.text_input("", placeholder="Ex: 6 meses", key="eact", label_visibility="collapsed")
    c1, c2, c3 = st.columns([2, 1, 2])
    c1.write("Estabilidade de longa duração")
    _l, _m, _r = c2.columns([1, 2, 1])
    estab_ld_aplic = _m.checkbox("", key="eld", label_visibility="collapsed")
    estab_ld_tempo = c3.text_input("", placeholder="Ex: 18 meses", key="eldt", label_visibility="collapsed")

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_6", use_container_width=True):
        st.session_state.next_tab = 6
        st.rerun()

# ── Aba 7 — Documentos ────────────────────────────────────────────────────────
with t7:
    step_info(7, "Documentos PDF", proxima="8. Embalagens")

    sec("13-18. DOCUMENTOS PDF")
    st.caption("Os PDFs serão incluídos junto com o arquivo Excel no ZIP.")
    pdfs = {}
    for nome, *_ in DOCS_INFO:
        with st.container(border=True):
            pdfs[nome] = st.file_uploader(nome, type=["pdf"], key=f"pdf_{nome}")

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_7", use_container_width=True):
        st.session_state.next_tab = 7
        st.rerun()

# ── Aba 8 — Embalagens ────────────────────────────────────────────────────────
with t8:
    step_info(8, "Embalagens e paletização — Seções 20 a 28", proxima="9. Fotos")

    _EMB_OPTS  = ["", "Bag 1,5L", "Bag 3L", "Frasco 1L", "Galão 5L", "Bombona 20L"]
    _T26_OPTS  = ["", "Offset", "Caixa Parda com Etiqueta"]

    qtd_apres = st.radio("Quantas apresentações este produto tem?", [1, 2],
                         horizontal=True, key="qtd_apres")

    # Dados de paletização padrão por embalagem
    _PALET_DADOS = {
        "Bag 1,5L":    {"lastro":"9",  "altura_cx":"6", "total_cx":"54", "total_lt":"648", "largura_mm":"1200", "altura_mm":"1335"},
        "Bag 3L":      {"lastro":"12", "altura_cx":"6", "total_cx":"72", "total_lt":"864", "largura_mm":"1200", "altura_mm":"1550"},
        "Frasco 1L":   {"lastro":"10", "altura_cx":"5", "total_cx":"50", "total_lt":"600", "largura_mm":"1200", "altura_mm":"1445"},
        "Galão 5L":    {"lastro":"10", "altura_cx":"4", "total_cx":"40", "total_lt":"800", "largura_mm":"1200", "altura_mm":"1520"},
        "Bombona 20L": {"lastro":"15", "altura_cx":"3", "total_cx":"45", "total_lt":"900", "largura_mm":"1200", "altura_mm":"1240"},
    }

    _apres_dados = {}

    for _idx, _suf in enumerate(["a", "b"][:qtd_apres]):
        _letra = _suf.upper()
        sec(f"APRESENTAÇÃO {_letra} — Embalagem Interna")

        _emb = st.selectbox(f"Embalagem interna {_letra}", _EMB_OPTS, key=f"tipo_emb_{_suf}")

        if _emb:
            _bk_apres = carregar_backend(_emb)
            _cx = _bk_apres.get("cx_int_nome", "")
            sec(f"Apresentação {_letra} — Caixa Interna")
            if _cx and _cx.upper() not in ("N/A", "NÃO APLICÁVEL"):
                st.info(f"Caixa interna: **{_cx}**")
            else:
                st.caption("Não aplicável para esta embalagem.")
            _qtd = _bk_apres.get("cx_mst_qtd", "")
            sec(f"Apresentação {_letra} — Caixa Master")
            if _qtd:
                st.info(f"Configuração: **{_qtd}**")
            sec(f"Apresentação {_letra} — Impressão da Caixa Master")
            _t26_sug = _bk_apres.get("t26_tipo", "")
            _t26_idx = _T26_OPTS.index(_t26_sug) if _t26_sug in _T26_OPTS else 0
            _t26_sel = st.selectbox("Tipo de impressão", _T26_OPTS, index=_t26_idx, key=f"t26_{_suf}")
            _bk_apres["t26_tipo"] = _t26_sel

            sec(f"Apresentação {_letra} — Caixa Master ERP / Etiqueta ERP (preenchimento manual)")
            st.caption("Esses campos não são preenchidos automaticamente. Preencha os valores manualmente.")
            _lbl2 = _bk_apres.get("t25_dim2", "Caixa Master ERP")
            _lbl3 = _bk_apres.get("t25_dim3", "Etiqueta ERP")
            _t25_med2 = st.text_input(_lbl2, key=f"t25_med2_{_suf}")
            _t25_med3 = st.text_input(_lbl3, key=f"t25_med3_{_suf}")
        else:
            _bk_apres = {}
            _t26_sel  = ""
            _t25_med2 = _t25_med3 = ""

        sec(f"Apresentação {_letra} — Paletização")
        _tp = st.radio("Tipo de paletização", ["Padrão", "Personalizada"],
                       horizontal=True, key=f"tipo_palet_{_suf}")
        _pd = _PALET_DADOS.get(_emb, {}) if (_tp == "Padrão" and _emb) else {}

        if _tp == "Padrão" and _emb:
            st.info(f"Valores preenchidos automaticamente para **{_emb}**.")

        c1, c2, c3 = st.columns(3)
        _p_lastro    = c1.text_input("Lastro (cx)",      value=_pd.get("lastro",""),    key=f"p_lastro_{_suf}",    disabled=(_tp=="Padrão"))
        _p_altura_cx = c2.text_input("Altura (cx)",      value=_pd.get("altura_cx",""), key=f"p_altura_cx_{_suf}", disabled=(_tp=="Padrão"))
        _p_total_cx  = c3.text_input("Total caixas",     value=_pd.get("total_cx",""),  key=f"p_total_cx_{_suf}",  disabled=(_tp=="Padrão"))
        c1, c2, c3 = st.columns(3)
        _p_total_lt  = c1.text_input("Total litros",     value=_pd.get("total_lt",""),  key=f"p_total_lt_{_suf}",  disabled=(_tp=="Padrão"))
        _p_largura   = c2.text_input("Largura (mm)",     value=_pd.get("largura_mm",""),key=f"p_largura_{_suf}",   disabled=(_tp=="Padrão"))
        _p_altura_mm = c3.text_input("Altura (mm)",      value=_pd.get("altura_mm",""), key=f"p_altura_mm_{_suf}", disabled=(_tp=="Padrão"))

        _bk_apres["tipo_emb_orig"] = _emb
        _apres_dados[_suf] = {
            "bk": _bk_apres, "t26": _t26_sel, "tipo_palet": _tp,
            "lastro": _pd.get("lastro","") if _tp=="Padrão" else _p_lastro,
            "altura_cx": _pd.get("altura_cx","") if _tp=="Padrão" else _p_altura_cx,
            "total_cx": _pd.get("total_cx","") if _tp=="Padrão" else _p_total_cx,
            "total_lt": _pd.get("total_lt","") if _tp=="Padrão" else _p_total_lt,
            "largura_mm": _pd.get("largura_mm","") if _tp=="Padrão" else _p_largura,
            "altura_mm": _pd.get("altura_mm","") if _tp=="Padrão" else _p_altura_mm,
            "t25_med2": _t25_med2, "t25_med3": _t25_med3,
        }

    # Expor variáveis da apresentação A para compatibilidade com código legado
    _a = _apres_dados.get("a", {})
    tipo_emb      = _a.get("bk", {}).get("tipo_emb_orig", "")
    _backend_dados = _a.get("bk", {})
    t26_tipo_sel  = _a.get("t26", "")
    tipo_palet    = _a.get("tipo_palet", "Padrão")
    palet_padrao  = _a.get("palet_padrao", "")
    palet_custom  = _a.get("palet_custom", "")

    # Variáveis legadas
    vol_int = ""; qtd_emb = ""; tipo_cx = ""; vol_tot = ""; palet = ""

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_8", use_container_width=True):
        st.session_state.next_tab = 8
        st.rerun()

# ── Aba 9 — Fotos ─────────────────────────────────────────────────────────────
with t9:
    step_info(9, "Fotos das embalagens — Seção 29", proxima="10. Assinatura")

    sec("29. IMAGENS DAS EMBALAGENS")
    st.caption("Serão centralizadas na ficha técnica.")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Caixa Master**")
        img_cx_master = st.file_uploader("Imagem Caixa Master",
            type=["png","jpg","jpeg","bmp","tiff"], key="img_cx_master")
    with c2:
        st.markdown("**Caixa Interna**")
        img_cx_interna = st.file_uploader("Imagem Caixa Interna",
            type=["png","jpg","jpeg","bmp","tiff"], key="img_cx_interna")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Frasco**")
        img_frasco = st.file_uploader("Imagem Frasco",
            type=["png","jpg","jpeg","bmp","tiff"], key="img_frasco")
    with c2:
        st.markdown("**Bag**")
        img_bag = st.file_uploader("Imagem Bag",
            type=["png","jpg","jpeg","bmp","tiff"], key="img_bag")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Galão**")
        img_galao = st.file_uploader("Imagem Galão",
            type=["png","jpg","jpeg","bmp","tiff"], key="img_galao")
    with c2:
        st.markdown("**Bombona**")
        img_bombona = st.file_uploader("Imagem Bombona",
            type=["png","jpg","jpeg","bmp","tiff"], key="img_bombona")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Etiqueta**")
        img_etiqueta123 = st.file_uploader("Imagem Etiqueta",
            type=["png","jpg","jpeg","bmp","tiff"], key="img_etiqueta123")

    _, col_btn = st.columns([3, 1])
    if col_btn.button("Próxima Página →", key="next_9", use_container_width=True):
        st.session_state.next_tab = 9
        st.rerun()

# ── Aba 10 — Assinatura ───────────────────────────────────────────────────────
with t10:
    step_info(10, "Assinatura do responsável")

    sec("ASSINATURA DO RESPONSÁVEL")
    nome_preenchido = st.text_input("Nome completo", key="nome_preenchido",
                                     placeholder="Nome de quem preencheu a ficha")
    data_hoje = date.today().strftime("%d/%m/%Y")
    st.caption(f"Data será preenchida automaticamente: **{data_hoje}**")

    # Variáveis legadas (dimensional antigo — não usadas no novo template)
    dim_frasco = False; dim_bag35 = False; dim_bag10 = False; dim_tampa = False
    emb_dim = ""; emb_mat = ""; emb_imp = ""; emb_faca = ""; emb_erp = ""
    cx_dim  = ""; cx_mat  = ""; cx_imp  = ""; cx_faca  = ""; cx_erp  = ""

    st.divider()
    gerar = st.button("⬇ Gerar Ficha Técnica", type="primary", use_container_width=True)

    if gerar:
        if not os.path.exists(TEMPLATE_PATH):
            st.error(f"Arquivo '{TEMPLATE_PATH}' não encontrado no repositório.")
            st.stop()

        try:
            wb = openpyxl.load_workbook(TEMPLATE_PATH)
            ws = wb.worksheets[0]

            # ── Montar dicionário de placeholders → valores ──────────────
            ph = {}

            # Cabeçalho (repetido em todas as páginas)
            ph["data"] = data
            ph["cod_ficha"] = cod_curto

            # Seção 1
            ph["cod_pa"]      = cod_pa
            ph["cod_pa_2"]    = cod_pa_2
            ph["cod_pa_3"]    = cod_pa_3
            ph["cod_pa_ppt1"] = cod_pa_ppt1
            ph["cod_pa_ppt2"] = cod_pa_ppt2
            ph["cod_int"]     = cod_int
            ph["cod_int_titulo"] = cod_int  # mesmo conteúdo, font 20 negrito
            ph["marca"] = marca
            ph["n_reg_1"] = n_reg_1

            # Seção 2 (até 6 micros)
            for i, m in enumerate(micros[:6]):
                n = i + 1
                ph[f"ing_{n}"] = m["ingrediente"]
                ph[f"cepa_{n}"] = m["cepa"]
                ph[f"cmin_{n}"] = m["conc_min"]
                ph[f"uni_{n}"] = m["unidade"]
                ph[f"gl_{n}"] = m.get("g_l", "")
                ph[f"mv_{n}"] = m.get("m_v", "")

            # Seção 3
            ph["classe"] = classe
            ph["shelf"] = shelf
            ph["formulac"] = formulac
            ph["armaz"] = armaz
            ph["via_reg"] = via_reg
            ph["ncm"] = ncm

            # Seção 4 e 5
            ph["titular"] = titular
            ph["info_tec"] = info_tec

            # Seção 20-26 — Embalagens (BACKEND)
            # Placeholders SEM sufixo (_a/_b/_c) → não preencher → mostram "-" no Excel
            # Placeholders COM sufixo são preenchidos pela lógica de apresentações (A/B/C)
            _bk = _backend_dados if tipo_emb else {}
            _T23_LABELS = ["Dimensão", "Caixa Interna ERP", "Material", "Impressão", "Acabamento"]
            _T25_LABELS = ["Dimensão", "Caixa Master ERP", "Etiqueta ERP", "Material", "Impressão"]
            _t26 = _bk.get("t26_tipo", "")

            def _preencher_apres(sufixo, bk, t26_sel):
                s  = f"_{sufixo}"
                emb = bk.get("tipo_emb_orig", "")
                ph[f"tipo_emb{s}"]    = emb
                ph[f"cx_int_nome{s}"] = bk.get("cx_int_nome", "")
                ph[f"cx_mst_qtd{s}"]  = bk.get("cx_mst_qtd", "")
                ph[f"cx_mst_nome{s}"] = bk.get("cx_mst_qtd", "")
                ph[f"t26_tipo{s}"]    = t26_sel

                # T21 — Embalagem Interna (nomes e valores do BACKEND)
                for i in range(1, 7):
                    ph[f"t21_dim{i}{s}"] = bk.get(f"t21_dim{i}", "")
                    ph[f"t21_med{i}{s}"] = bk.get(f"t21_med{i}", "")

                # T23 — Caixa Interna (labels fixos, valores do BACKEND ou N/A)
                t23_labels = ["Dimensão", "Código ERP Cx.I.", "Material", "Impressão", "Acabamento"]
                t23_na = bk.get("t23_dim1", "") in ("N/A", "")
                for i, lbl in enumerate(t23_labels, start=1):
                    ph[f"t23_dim{i}{s}"] = lbl
                    ph[f"t23_med{i}{s}"] = "N/A" if t23_na else (bk.get(f"t23_med{i}") or "")
                ph[f"t23_dim6{s}"] = ""
                ph[f"t23_med6{s}"] = ""

                # T25 — Caixa Master (labels e valores do BACKEND)
                for i in range(1, 7):
                    ph[f"t25_dim{i}{s}"] = bk.get(f"t25_dim{i}") or ""
                    if i == 3:  # Etiqueta ERP — condicional ao T26
                        ph[f"t25_med{i}{s}"] = "N/A" if t26_sel == "Offset" else (bk.get(f"t25_med{i}") or "")
                    else:
                        ph[f"t25_med{i}{s}"] = bk.get(f"t25_med{i}") or ""

            # Preencher placeholders para cada apresentação selecionada
            for _suf_ap, _ap_info in _apres_dados.items():
                _bk_ap  = _ap_info.get("bk", {})
                _t26_ap = _ap_info.get("t26", "")
                _preencher_apres(_suf_ap, _bk_ap, _t26_ap)
                _s = f"_{_suf_ap}"
                # t25_med2 e t25_med3 — preenchidos manualmente (sobrescreve BACKEND)
                ph[f"t25_med2{_s}"] = _ap_info.get("t25_med2", "")
                ph[f"t25_med3{_s}"] = _ap_info.get("t25_med3", "")
                # Paletização — campos individuais
                ph[f"palet_lastro{_s}"]    = _ap_info.get("lastro", "")
                ph[f"palet_altura_cx{_s}"] = _ap_info.get("altura_cx", "")
                ph[f"palet_total_cx{_s}"]  = _ap_info.get("total_cx", "")
                ph[f"palet_total_lt{_s}"]  = _ap_info.get("total_lt", "")
                ph[f"palet_largura_mm{_s}"]= _ap_info.get("largura_mm", "")
                ph[f"palet_altura_mm{_s}"] = _ap_info.get("altura_mm", "")
            # Placeholders antigos sem sufixo → não preencher → mostram "-"
            # Campos legados mantidos vazios
            ph["vol_int"] = vol_int
            ph["qtd_emb"] = qtd_emb
            ph["tipo_cx"] = tipo_cx
            ph["vol_tot"] = vol_tot
            ph["palet"]   = palet

            # Seção 7 — Recomendações
            for i, rv in enumerate(rec_rows):
                n = i + 1
                ph[f"cult_{n}"] = "Todas as culturas com ocorrência do alvo biológico"
                ph[f"alvo_{n}"] = rv.get("alvos", "")
                ph[f"dose_{n}"] = rv.get("dose", "")
                ph[f"vol_{n}"] = rv.get("volume", "")
                ph[f"num_{n}"] = rv.get("numero", "")

            # Seção 8 — Qualidade (micros 1-6)
            _q_campos = ["ingrediente","estado_fisico","cor","aspecto",
                         "gram","crescimento","meio","morfologia"]
            _q_keys   = ["ing","est","cor","asp","gram","cresc","meio","morf"]
            for i, vq in enumerate(qualidade):
                n = i + 1
                for key, campo in zip(_q_keys, _q_campos):
                    ph[f"q_{key}_{n}"] = vq.get(campo, "")

            # Seção 9 — Análises
            _an_keys = ["pureza","conc","ph","dens","visc","teor"]
            for (aplic, vmin, vmax), akey in zip(analises, _an_keys):
                ph[f"ap_{akey}"] = sn(aplic)
                ph[f"min_{akey}"] = vmin
                ph[f"max_{akey}"] = vmax
            ph["unidade_conc"] = unidade_conc

            # Seção 11 — descrições das figuras
            for n in range(1, 7):
                ph[f"desc_fig{n}"] = descs_fig.get(n, "")

            # Seção 12 — CoA
            ph["coa_i"] = sn(coa_i)
            ph["coa_ii"] = sn(coa_ii)
            # CoA III — análises selecionadas
            # coa_iii_2 (Análise de concentração) removido do formulário — sempre "-"
            ph["coa_iii_2"] = "-"
            _coa_iii_names = ["Análise de Pureza", "PH", "Densidade",
                              "Viscosidade", "Teor de umidade/água/granulometria"]
            _coa_iii_slots = [1, 3, 4, 5, 6]  # slot 2 reservado para o hífen fixo
            for slot, nome_c in zip(_coa_iii_slots, _coa_iii_names):
                ph[f"coa_iii_{slot}"] = nome_c if ('coa_iii_selecionadas' in dir() and nome_c in coa_iii_selecionadas) else ""
            ph["coa_iv"] = sn(coa_iv)
            ph["justif"] = justif

            # Seção 14-18 — Documentos
            _doc_map = {"14. Bula": "bula", "15. Rótulo": "rotulo",
                        "16. Certificado de Registro": "cert",
                        "17. Ficha de Emergência": "ficha",
                        "18. FDS": "fds", "18. SDS": "sds",
                        "Estudos de Estabilidade Acelerada": "acelerado",
                        "Estudos de Longa Duração": "longa_duracao"}
            for nome_doc, suf in _doc_map.items():
                _pdf_nome = f"{suf.upper()}.pdf"
                ph[f"pdf_{suf}"]  = _pdf_nome if pdfs.get(nome_doc) else ""
                ph[f"data_{suf}"] = data if pdfs.get(nome_doc) else ""

            # Seção 19
            ph["estab_ac_ap"] = sn(estab_ac_aplic)
            ph["estab_ac_t"] = estab_ac_tempo
            ph["estab_ld_ap"] = sn(estab_ld_aplic)
            ph["estab_ld_t"] = estab_ld_tempo

            # Seção 20
            ph["dim_frasco"] = sn(dim_frasco)
            ph["dim_bag35"] = sn(dim_bag35)
            ph["dim_bag10"] = sn(dim_bag10)
            ph["dim_tampa"] = sn(dim_tampa)
            ph["emb_dim"] = emb_dim
            ph["emb_mat"] = emb_mat
            ph["emb_imp"] = emb_imp
            ph["emb_faca"] = emb_faca
            ph["emb_erp"] = emb_erp
            ph["cx_dim"] = cx_dim
            ph["cx_mat"] = cx_mat
            ph["cx_imp"] = cx_imp
            ph["cx_faca"] = cx_faca
            ph["cx_erp"] = cx_erp

            # Seção 23
            ph["nome_preenchido"] = nome_preenchido
            ph["data_preenchido"] = data

            # ── Buscar e substituir placeholders no template ─────────────
            import re as _re
            _ph_pattern = _re.compile(r'\{\{([^}]+)\}\}')

            # Mapear placeholders de imagem (não substituir por texto)
            _img_placeholders = {
                "img_fig1","img_fig2","img_fig3","img_fig4","img_fig5","img_fig6",
                "img_coa","img_cx_master","img_cx_interna","img_frasco","img_bag",
                "img_galao","img_bombona","img_etiqueta123"
            }

            # Registrar células de imagem para inserção posterior
            img_cells = {}

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "{{" in cell.value:
                        match = _ph_pattern.search(cell.value)
                        if match:
                            key = match.group(1)
                            if key in _img_placeholders:
                                img_cells[key] = cell.coordinate
                                cell.value = None  # Limpar placeholder de imagem
                            elif key in ph and ph[key]:
                                val = ph[key]
                                # PDFs: criar hyperlink clicável
                                if key.startswith("pdf_") and val.endswith(".pdf"):
                                    cell.value = val
                                    cell.hyperlink = val
                                    cell.font = Font(color="0563C1", underline="single")
                                elif key == "cod_int_titulo":
                                    cell.value = val
                                    cell.font = Font(color="000000", size=20, bold=True)
                                    cell.alignment = Alignment(horizontal="left",
                                                               vertical="center",
                                                               wrap_text=True)
                                    continue
                                else:
                                    cell.value = val
                                    cell.font = Font(color="000000")
                                # Alinhamento por tópico
                                _left_keys = {"cod_pa","cod_int","marca","titular","info_tec"}
                                _left_pfx  = ("q_ing_","q_est_","q_cor_","q_asp_","q_gram_","q_cresc_","q_meio_","q_morf_","desc_fig")
                                _h = "left" if (key in _left_keys or any(key.startswith(p) for p in _left_pfx)) else "center"
                                _v = "top" if key.startswith("desc_fig") else "center"
                                cell.alignment = Alignment(horizontal=_h,
                                                           vertical=_v,
                                                           wrap_text=True)
                                # Formatação condicional SIM/NÃO
                                if val == "SIM":
                                    cell.fill = PatternFill(start_color="C6EFCE",
                                                            end_color="C6EFCE", fill_type="solid")
                                elif val == "NÃO":
                                    cell.fill = PatternFill(start_color="FFC7CE",
                                                            end_color="FFC7CE", fill_type="solid")
                            else:
                                # Placeholder não preenchido → hífen com fundo #F2F2F2
                                cell.value = "-"
                                cell.font = Font(color="000000")
                                cell.alignment = Alignment(horizontal="center",
                                                           vertical="center")
                                cell.fill = PatternFill(start_color="F2F2F2",
                                                        end_color="F2F2F2", fill_type="solid")

            # ── Inserir imagens ──────────────────────────────────────────
            imgs_para_zip = {}

            # Dimensões exatas das células por placeholder (px)
            _cell_dims = {
                "img_fig1": (880, 200), "img_fig2": (880, 200), "img_fig3": (880, 180),
                "img_fig4": (880, 200), "img_fig5": (880, 200), "img_fig6": (880, 180),
                "img_coa":  (880, 860), "palet_padrao": (880, 320),
                "img_cx_master": (425, 140), "img_cx_interna": (455, 140),
                "img_frasco":    (425, 140), "img_bag":        (455, 140),
                "img_galao":     (455, 140), "img_bombona":    (455, 140),
                "img_etiqueta123": (455, 140),
            }

            # Seção 11 — figuras 1-6
            for n in range(1, 7):
                up = imgs_fig.get(n)
                anchor_key = f"img_fig{n}"
                if up and anchor_key in img_cells:
                    _w, _h = _cell_dims.get(anchor_key, (640, 140))
                    inserir_imagem_centralizada(ws, up.getvalue(),
                                                img_cells[anchor_key], _w, _h)
                    imgs_para_zip[f"Figura {n} - {up.name}"] = up.getvalue()

            # Seção 13 — Modelo CoA
            if "img_coa" in img_cells:
                for ext in ("png", "jpg", "jpeg", "bmp", "tiff"):
                    modeloca_path = f"MODELOCA.{ext}"
                    if os.path.exists(modeloca_path):
                        _w, _h = _cell_dims.get("img_coa", (880, 860))
                        inserir_imagem_centralizada(ws, modeloca_path,
                                                    img_cells["img_coa"], _w, _h)
                        break

            # Seção 21 — embalagens
            _emb_map = {
                "img_cx_master":  img_cx_master,
                "img_cx_interna": img_cx_interna,
                "img_frasco":     img_frasco,
                "img_bag":        img_bag,
                "img_galao":        img_galao,
                "img_bombona":      img_bombona,
                "img_etiqueta123":  img_etiqueta123,
            }
            for key, up_emb in _emb_map.items():
                if up_emb and key in img_cells:
                    _w, _h = _cell_dims.get(key, (425, 140))
                    inserir_imagem_centralizada(ws, up_emb.getvalue(),
                                                img_cells[key], _w, _h)
                    nome_emb = key.replace("img_", "").replace("_", " ").title()
                    imgs_para_zip[f"Embalagem - {nome_emb} - {up_emb.name}"] = up_emb.getvalue()

            # Configurar página A4 antes de salvar
            ws.page_setup.paperSize = 9  # 9 = A4 (padrão ECMA-376)

            # Salvar Excel em memória
            excel_buf = io.BytesIO()
            wb.save(excel_buf)
            excel_bytes = excel_buf.getvalue()

            # Montar ZIP
            cod_limpo  = re.sub(r'[<>:"/\\|?*]', "-", cod_int).strip() or "Ficha_Tecnica"
            _cod_int_limpo   = re.sub(r'[<>:"/\\|?*]', "-", cod_int).strip()  or "INT"
            _cod_ficha_limpo = re.sub(r'[<>:"/\\|?*]', "-", cod_curto).strip() or "FICHA"
            nome_excel = f"ATLAS - {_cod_int_limpo} - {_cod_ficha_limpo}.xlsx"
            zip_buf    = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(nome_excel, excel_bytes)
                for nome_doc, _, __, nome_pdf in DOCS_INFO:
                    if pdfs.get(nome_doc):
                        zf.writestr(f"{nome_pdf}.pdf", pdfs[nome_doc].getvalue())
                for nome_img, img_bytes in imgs_para_zip.items():
                    zf.writestr(nome_img, img_bytes)
            zip_buf.seek(0)

            st.success("✅ Ficha gerada com sucesso!")
            st.download_button(
                label="⬇ Baixar Ficha + Documentos (.zip)",
                data=zip_buf,
                file_name=f"{cod_limpo}.zip",
                mime="application/zip",
                use_container_width=True,
            )

        except Exception as e:
            import traceback
            st.error(f"Erro ao gerar: {e}")
            with st.expander("Detalhes do erro"):
                st.code(traceback.format_exc())

# ── Injeção JS para navegação entre abas ──────────────────────────────────────
if st.session_state.next_tab is not None:
    idx = st.session_state.next_tab
    st.session_state.next_tab = None
    components.html(f"""
        <script>
            setTimeout(function() {{
                var tabs = window.parent.document.querySelectorAll('button[role="tab"]');
                if (tabs && tabs[{idx}]) tabs[{idx}].click();
            }}, 150);
        </script>
    """, height=0)
